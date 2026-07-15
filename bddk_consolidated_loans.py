#!/usr/bin/env python3
"""
BDDK Consolidated Bank Report Scraper — Consumer Loan Breakdown
=================================================================

Downloads consolidated ("Konsolide") independent audit reports for every
Turkish bank from BDDK's "Bağımsız Denetim Raporları" portal
(https://www.bddk.gov.tr/BdrUyg), extracts the mandatory "Tüketici
kredileri ... ilişkin bilgiler" footnote table from each report's PDF,
and writes the results into an Excel workbook that follows the layout of
Template.xlsx (one block of rows per bank, one column per quarter).

Fields extracted per bank per quarter (all in thousand TRY, as reported):
    Tüketici Kredileri-TP  (Toplam)
        Konut Kredisi
        Taşıt Kredisi
        İhtiyaç Kredisi
        Diğer
    Tüketici Kredileri-YP  (Toplam)
        Konut Kredisi
        Taşıt Kredisi
        İhtiyaç Kredisi
        Diğer

Rules (per user spec):
    - If a bank has no consolidated report for a given quarter -> all 10
      fields for that bank/quarter are 0.
    - If a field is blank / "-" in the report -> 0.

Usage:
    python3 bddk_consolidated_loans.py
    python3 bddk_consolidated_loans.py --banks 5          # limit for a quick test
    python3 bddk_consolidated_loans.py --start 2023Q1 --end 2026Q1
    python3 bddk_consolidated_loans.py --output result.xlsx --cache-dir ./bddk_cache

Notes:
    - Downloaded ZIP/PDF files are cached under --cache-dir so re-running
      the script (e.g. after an interruption) does not re-download data
      that was already fetched.
    - A full run covers ~70 banks x up to 13 quarters. Most banks will
      have no consolidated report (only banking groups with subsidiaries
      file one), but the script still queries every bank/quarter. Expect
      the full run to take a while (network + PDF parsing bound); use
      --banks / --start / --end to do a smaller test run first.
"""

import argparse
import io
import json
import os
import re
import sys
import time
import urllib.parse
import zipfile
from dataclasses import dataclass, field

import docx
import openpyxl
import pdfplumber
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    # bddk.gov.tr serves an incomplete certificate chain (missing
    # intermediate CA); the OS trust store can resolve it via AIA
    # fetching but certifi's static bundle cannot. truststore routes
    # verification through the OS trust store instead.
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

BASE = "https://www.bddk.gov.tr/BdrUyg"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )
}

FIELD_ORDER = [
    ("TP_Toplam", "Tüketici Kredileri-TP"),
    ("TP_Konut", "Konut Kredisi"),
    ("TP_Tasit", "Taşıt Kredisi"),
    ("TP_Ihtiyac", "İhtiyaç Kredisi"),
    ("TP_Diger", "Diğer"),
    ("YP_Toplam", "Tüketici Kredileri-YP"),
    ("YP_Konut", "Konut Kredisi"),
    ("YP_Tasit", "Taşıt Kredisi"),
    ("YP_Ihtiyac", "İhtiyaç Kredisi"),
    ("YP_Diger", "Diğer"),
]

ZERO_RESULT = {key: 0 for key, _ in FIELD_ORDER}

# Verified manual overrides, keyed by (eft_kodu, year, month).
#
# These three consolidated reports render the "Tüketici kredileri ..." table
# in a way the PDF text layer cannot express (Fibabanka 2023Q3: per-glyph
# raster images; Fibabanka 2024Q1: vector curves; TFKB 2026Q1: letter-spaced
# text with value columns dumped separately). The pages were rasterized
# (see bddk_cache/_overrides/*.png) and the 10 template fields read directly
# from the image. Each block is self-checking: the four sub-items
# (Konut + Taşıt + İhtiyaç + Diğer) sum exactly to the -TP / -YP total, which
# assert_overrides_valid() re-verifies at startup so a bad transcription
# fails loudly. Values are in thousand TRY (Toplam column), matching the
# normal text-parse output.
MANUAL_OVERRIDES = {
    # FİBABANKA A.Ş. (EFT 103) — 30.09.2023, p.60 (rendered p60)
    (103, 2023, 9): {
        "TP_Toplam": 11119514, "TP_Konut": 81896, "TP_Tasit": 498,
        "TP_Ihtiyac": 11037120, "TP_Diger": 0,
        "YP_Toplam": 0, "YP_Konut": 0, "YP_Tasit": 0, "YP_Ihtiyac": 0, "YP_Diger": 0,
    },
    # FİBABANKA A.Ş. (EFT 103) — 31.03.2024, p.56 (rendered p56)
    (103, 2024, 3): {
        "TP_Toplam": 11806193, "TP_Konut": 67580, "TP_Tasit": 484,
        "TP_Ihtiyac": 11738129, "TP_Diger": 0,
        "YP_Toplam": 0, "YP_Konut": 0, "YP_Tasit": 0, "YP_Ihtiyac": 0, "YP_Diger": 0,
    },
    # TÜRKİYE FİNANS KATILIM BANKASI A.Ş. (EFT 206) — 31.03.2026, p.52 (rendered p52)
    (206, 2026, 3): {
        "TP_Toplam": 15012676, "TP_Konut": 4589164, "TP_Tasit": 1597563,
        "TP_Ihtiyac": 8825949, "TP_Diger": 0,
        "YP_Toplam": 0, "YP_Konut": 0, "YP_Tasit": 0, "YP_Ihtiyac": 0, "YP_Diger": 0,
    },
}


def assert_overrides_valid():
    """Re-check every manual override's internal consistency at startup.

    Guards against a typo in MANUAL_OVERRIDES: the four consumer-loan
    sub-items must sum to the corresponding -TP / -YP total.
    """
    for (eft, year, month), vals in MANUAL_OVERRIDES.items():
        for cur in ("TP", "YP"):
            total = vals[f"{cur}_Toplam"]
            parts = sum(vals[f"{cur}_{k}"] for k in ("Konut", "Tasit", "Ihtiyac", "Diger"))
            if total != parts:
                raise AssertionError(
                    f"Override {(eft, year, month)} {cur}: sub-items sum to "
                    f"{parts} but total is {total}"
                )


# The 25 banks that actually file consolidated consumer-loan data, in the
# order requested for the filtered output workbook. Names must match the
# BankaAdi values returned by KurulusListesiGetir exactly.
WANTED_BANKS = [
    "AKBANK T.A.Ş.",
    "ALBARAKA TÜRK KATILIM BANKASI A.Ş.",
    "ALTERNATİFBANK A.Ş.",
    "ANADOLUBANK A.Ş.",
    "BURGAN BANK A.Ş.",
    "DENİZBANK A.Ş.",
    "FİBABANKA A.Ş.",
    "HAYAT FİNANS KATILIM BANKASI A.Ş.",
    "HSBC BANK A.Ş.",
    "ICBC TURKEY BANK A.Ş.",
    "ING BANK A.Ş.",
    "KUVEYT TÜRK KATILIM BANKASI A.Ş.",
    "QNB BANK A.Ş.",
    "T.C. ZİRAAT BANKASI A.Ş.",
    "TÜRK EKONOMİ BANKASI A.Ş.",
    "TÜRKİYE EMLAK KATILIM BANKASI A.Ş.",
    "TÜRKİYE FİNANS KATILIM BANKASI A.Ş.",
    "TÜRKİYE GARANTİ BANKASI A.Ş.",
    "TÜRKİYE HALK BANKASI A.Ş.",
    "TÜRKİYE VAKIFLAR BANKASI T.A.O.",
    "TÜRKİYE İŞ BANKASI A.Ş.",
    "VAKIF KATILIM BANKASI A.Ş.",
    "YAPI VE KREDİ BANKASI A.Ş.",
    "ZİRAAT KATILIM BANKASI A.Ş.",
    "ŞEKERBANK T.A.Ş.",
]

# Case-insensitive: some banks' PDFs use lowercase table labels (e.g. Yapı
# Kredi's "Tüketici kredileri-TP" vs Garanti's "Tüketici Kredileri-TP").
# Dash class covers '-', en-dash '–', em-dash '—' seen across different
# banks' report templates (e.g. ING uses "kredileri – TP").
# Word-gap is \s* (not \s+): some banks' PDFs have tightly-kerned text
# where pdfplumber extracts no space at all between words (e.g. Albaraka
# Türk's "TüketiciKredileri-TP", "KonutKredisi").
_DASH = r"[-–—]"
TP_HEADER_RE = re.compile(r"^T[üu]ketici\s*Kredileri\s*" + _DASH + r"\s*TP\b", re.IGNORECASE)
YP_HEADER_RE = re.compile(r"^T[üu]ketici\s*Kredileri\s*" + _DASH + r"\s*YP\b", re.IGNORECASE)
KONUT_RE = re.compile(r"^Konut\s*Kredisi\b", re.IGNORECASE)
TASIT_RE = re.compile(r"^Ta[şs][ıi]t\s*Kredisi\b", re.IGNORECASE)
IHTIYAC_RE = re.compile(r"^[İIi]htiya[çc]\s*Kredisi\b", re.IGNORECASE)
DIGER_RE = re.compile(r"^Di[ğg]er\b", re.IGNORECASE)

NUM_TOKEN_RE = re.compile(r"\(?-?\d[\d,.]*\)?|(?<!\S)-(?!\S)")


def make_session():
    session = requests.Session()
    session.headers.update(HEADERS)
    retry = Retry(total=5, backoff_factor=1.5, status_forcelist=[429, 500, 502, 503, 504])
    session.mount("https://", HTTPAdapter(max_retries=retry))
    session.mount("http://", HTTPAdapter(max_retries=retry))
    return session


def normalize_line(line):
    return line.replace("\xa0", " ").strip()


def parse_number(tok):
    tok = tok.strip()
    if tok in ("-", "–", "—", ""):
        return 0
    neg = tok.startswith("(") and tok.endswith(")")
    if neg:
        tok = tok[1:-1]
    if re.fullmatch(r"-?\d{1,3}(,\d{3})+(\.\d+)?", tok):
        tok = tok.replace(",", "")
    elif re.fullmatch(r"-?\d{1,3}(\.\d{3})+(,\d+)?", tok):
        tok = tok.replace(".", "").replace(",", ".")
    try:
        val = float(tok)
    except ValueError:
        return 0
    val = -val if neg else val
    return int(val) if val == int(val) else val


def line_last_number(line):
    toks = NUM_TOKEN_RE.findall(line)
    if not toks:
        return None
    return parse_number(toks[-1])


def _pdf_to_lines(pdf_path):
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines.extend(normalize_line(l) for l in text.split("\n"))
    return lines


def _docx_to_lines(docx_path):
    # Flatten every table's rows, in document order, into "label val1 val2
    # val3" strings so the same line-based parser used for PDFs applies
    # unchanged. Row adjacency within a table is preserved by processing
    # tables/rows sequentially.
    doc = docx.Document(docx_path)
    lines = []
    for table in doc.tables:
        for row in table.rows:
            cells = [normalize_line(c.text) for c in row.cells]
            if cells and cells[0]:
                lines.append(" ".join(cells))
    return lines


def extract_consumer_loans(report_path):
    """Parse the 'Tüketici kredileri ... ilişkin bilgiler' footnote table.

    Returns (result_dict, status_string). result_dict always has all 10
    keys; any field that could not be located defaults to 0 per spec.
    """
    if report_path.lower().endswith(".docx"):
        lines = _docx_to_lines(report_path)
    else:
        lines = _pdf_to_lines(report_path)

    result = dict(ZERO_RESULT)
    n = len(lines)

    tp_idx = next((i for i, l in enumerate(lines) if TP_HEADER_RE.match(l)), None)
    if tp_idx is None:
        return result, "TP header not found (bank likely has no consumer-loan note)"

    result["TP_Toplam"] = line_last_number(lines[tp_idx]) or 0

    def find_after(start, pattern, window=8):
        for j in range(start + 1, min(start + 1 + window, n)):
            if pattern.match(lines[j]):
                return j
        return None

    idx = tp_idx
    for key, pattern in (("TP_Konut", KONUT_RE), ("TP_Tasit", TASIT_RE),
                          ("TP_Ihtiyac", IHTIYAC_RE), ("TP_Diger", DIGER_RE)):
        found = find_after(idx, pattern)
        if found is not None:
            result[key] = line_last_number(lines[found]) or 0
            idx = found

    yp_idx = None
    for j in range(idx + 1, min(idx + 1 + 20, n)):
        if YP_HEADER_RE.match(lines[j]):
            yp_idx = j
            break
    if yp_idx is None:
        return result, "YP header not found"

    result["YP_Toplam"] = line_last_number(lines[yp_idx]) or 0
    idx = yp_idx
    for key, pattern in (("YP_Konut", KONUT_RE), ("YP_Tasit", TASIT_RE),
                          ("YP_Ihtiyac", IHTIYAC_RE), ("YP_Diger", DIGER_RE)):
        found = find_after(idx, pattern)
        if found is not None:
            result[key] = line_last_number(lines[found]) or 0
            idx = found

    return result, "ok"


def get_banks(session):
    r = session.post(f"{BASE}/Home/KurulusListesiGetir", data={"kurulusTuruId": 1, "kurulusGrubuId": ""})
    r.raise_for_status()
    banks = r.json()
    return sorted(banks, key=lambda b: b["BankaAdi"])


RAPOR_URL_RE = re.compile(r'raporUrl=([^"&]+)')
KONSOLIDE_FILE_RE = re.compile(r"KONSOLIDE-(\d{4})-(\d{2})\.zip$", re.IGNORECASE)


def get_year_reports(session, eft_kodu, year, rapor_tipi="KONSOLIDE"):
    """Returns {month: raporUrl} for one bank/year via a single request."""
    params = {
        "KurulusTuru": 1,
        "EFTKodu": eft_kodu,
        "RaporTipi": rapor_tipi,
        "DonemYil": year,
        "DonemAy": 0,
    }
    r = session.get(f"{BASE}/Home/SorguSonuc", params=params)
    r.raise_for_status()
    out = {}
    for raw_url in RAPOR_URL_RE.findall(r.text):
        url = raw_url.replace("&amp;", "&")
        m = KONSOLIDE_FILE_RE.search(requests.utils.unquote(url))
        if m:
            out[int(m.group(2))] = url
    return out


def download_pdf(session, rapor_url, cache_dir, cache_key):
    """Downloads+extracts the report ZIP (cached), returns path to the PDF."""
    os.makedirs(cache_dir, exist_ok=True)
    pdf_dir = os.path.join(cache_dir, cache_key)
    marker = os.path.join(pdf_dir, ".pdf_path")
    if os.path.exists(marker):
        with open(marker) as f:
            p = f.read().strip()
        if p and os.path.exists(p):
            return p

    # rapor_url comes from the search-results HTML already percent-encoded
    # (e.g. "~%2FDosya%2F...zip"); requests.params would re-encode the '%'
    # signs and double-encode it, so decode first and let requests encode once.
    decoded_url = urllib.parse.unquote(rapor_url)
    r = session.get(f"{BASE}/Home/DosyaIndir", params={"raporUrl": decoded_url}, timeout=120)
    r.raise_for_status()
    os.makedirs(pdf_dir, exist_ok=True)

    if r.content[:4] == b"%PDF":
        # Some reports are served as a raw PDF despite the .zip filename/
        # application/zip content-type (e.g. HSBC).
        pdf_path = os.path.join(pdf_dir, "report.pdf")
        with open(pdf_path, "wb") as f:
            f.write(r.content)
    else:
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            z.extractall(pdf_dir)

        # Some archives are a zip-within-a-zip (e.g. Nurol, TEB): the outer
        # zip's only entry is itself a .zip, not a report file. Unwrap up
        # to a few levels until we find a PDF or DOCX.
        for _ in range(3):
            pdf_path = _find_report_file(pdf_dir)
            if pdf_path:
                break
            nested_zip = _find_ext(pdf_dir, ".zip")
            if not nested_zip:
                break
            with zipfile.ZipFile(nested_zip) as z:
                z.extractall(pdf_dir)
        pdf_path = _find_report_file(pdf_dir)

    if pdf_path is None:
        raise RuntimeError(f"No PDF/DOCX report found inside archive for {rapor_url}")

    with open(marker, "w") as f:
        f.write(pdf_path)
    return pdf_path


def _find_report_file(root_dir):
    # Some archives contain more than one PDF/DOCX (e.g. a short
    # "responsibility statement" alongside the full financial statements,
    # as with Nurol). The full financial-statement report is always by far
    # the largest file, so picking the largest candidate reliably selects
    # the right one. A handful of banks (e.g. TEB) file .docx instead of PDF.
    candidates = []
    for root, _dirs, files in os.walk(root_dir):
        for fn in files:
            if fn.lower().endswith((".pdf", ".docx")):
                path = os.path.join(root, fn)
                candidates.append((os.path.getsize(path), path))
    if not candidates:
        return None
    return max(candidates)[1]


def _find_ext(root_dir, ext):
    for root, _dirs, files in os.walk(root_dir):
        for fn in files:
            if fn.lower().endswith(ext):
                return os.path.join(root, fn)
    return None


def quarter_range(start, end):
    """start/end like '2023Q1' -> ordered list of (year, month, label)."""
    def parse(q):
        y, qn = int(q[:4]), int(q[5])
        return y, qn

    sy, sq = parse(start)
    ey, eq = parse(end)
    out = []
    y, q = sy, sq
    while (y, q) <= (ey, eq):
        out.append((y, q * 3, f"{y}Q{q}"))
        q += 1
        if q > 4:
            q = 1
            y += 1
    return out


# Layout/styling copied from Template.xlsx: labels in column F, quarters
# starting at column G, bold+yellow bank-name/header rows, plain field rows.
LABEL_COL = 6   # F
FIRST_Q_COL = 7  # G

HEADER_FONT = openpyxl.styles.Font(name="Helvetica", size=12, bold=True)
QUARTER_FONT = openpyxl.styles.Font(name="Helvetica", size=12, bold=False)
FIELD_FONT = openpyxl.styles.Font(name="Helvetica", size=12, bold=False)
YELLOW_FILL = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FFFFFF00")
VALUE_ALIGN = openpyxl.styles.Alignment(horizontal="left")


def build_workbook(quarters, bank_results, bank_order):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    row = 1
    first_bank = True
    for bank_name in bank_order:
        name_cell = ws.cell(row=row, column=LABEL_COL, value=bank_name)
        name_cell.font = HEADER_FONT
        name_cell.fill = YELLOW_FILL
        if first_bank:
            # Template fuses the quarter header labels into the first
            # bank's name row (e.g. F46='GARANTİ', G46='2023 Q1', ...).
            for j, (y, q3, _label) in enumerate(quarters):
                qcell = ws.cell(row=row, column=FIRST_Q_COL + j, value=f"{y} Q{q3 // 3}")
                qcell.font = QUARTER_FONT
                qcell.fill = YELLOW_FILL
            first_bank = False
        row += 1

        for key, field_label in FIELD_ORDER:
            label_cell = ws.cell(row=row, column=LABEL_COL, value=field_label)
            label_cell.font = FIELD_FONT
            for j, (_y, _m, qlabel) in enumerate(quarters):
                val = bank_results.get(bank_name, {}).get(qlabel, ZERO_RESULT)[key]
                vcell = ws.cell(row=row, column=FIRST_Q_COL + j, value=val)
                vcell.font = FIELD_FONT
                vcell.alignment = VALUE_ALIGN
                vcell.number_format = "#,##0"
            row += 1

    ws.column_dimensions[openpyxl.utils.get_column_letter(LABEL_COL)].width = 35.5
    for j in range(len(quarters)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(FIRST_Q_COL + j)].width = 17

    ws.freeze_panes = ws.cell(row=2, column=FIRST_Q_COL).coordinate

    return wb


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--start", default="2023Q1", help="First quarter, e.g. 2023Q1")
    ap.add_argument("--end", default="2026Q1", help="Last quarter, e.g. 2026Q1")
    ap.add_argument("--output", default="BDDK_Konsolide_Tuketici_Kredileri.xlsx")
    ap.add_argument("--cache-dir", default="bddk_cache")
    ap.add_argument("--banks", type=int, default=None, help="Limit number of banks (for testing)")
    ap.add_argument("--only-banks", action="store_true",
                    help="Restrict output to the 25 WANTED_BANKS, in that order.")
    ap.add_argument("--delay", type=float, default=0.4, help="Delay in seconds between HTTP requests")
    ap.add_argument("--log", default="bddk_scrape_log.jsonl")
    args = ap.parse_args()

    assert_overrides_valid()

    quarters = quarter_range(args.start, args.end)
    years_needed = sorted({y for y, _m, _l in quarters})

    session = make_session()

    print("Fetching bank list...")
    banks = get_banks(session)
    if args.only_banks:
        by_name = {b["BankaAdi"]: b for b in banks}
        missing = [n for n in WANTED_BANKS if n not in by_name]
        if missing:
            raise SystemExit("WANTED_BANKS not found in bank list: " + "; ".join(missing))
        banks = [by_name[n] for n in WANTED_BANKS]
    if args.banks:
        banks = banks[: args.banks]
    print(f"{len(banks)} banks to process, {len(quarters)} quarters each "
          f"({args.start} .. {args.end}).")

    bank_results = {}
    bank_order = [b["BankaAdi"] for b in banks]
    logf = open(args.log, "a", encoding="utf-8")

    for bi, bank in enumerate(banks, 1):
        eft = bank["EFTKodu"]
        name = bank["BankaAdi"]
        print(f"[{bi}/{len(banks)}] {name} (EFT {eft})")
        bank_results[name] = {}

        year_reports = {}
        for year in years_needed:
            try:
                year_reports[year] = get_year_reports(session, eft, year)
            except Exception as e:
                print(f"  WARN: could not fetch report index for {year}: {e}")
                year_reports[year] = {}
            time.sleep(args.delay)

        for year, month, qlabel in quarters:
            override = MANUAL_OVERRIDES.get((eft, year, month))
            if override is not None:
                # Verified figures read from the rendered page image for the
                # few reports whose table isn't in the PDF text layer.
                bank_results[name][qlabel] = dict(override)
                logf.write(json.dumps({"bank": name, "eft": eft, "quarter": qlabel,
                                        "status": "manual_override"}, ensure_ascii=False) + "\n")
                print(f"  OVERRIDE {qlabel}: using verified values from rendered image")
                continue

            rapor_url = year_reports.get(year, {}).get(month)
            if not rapor_url:
                bank_results[name][qlabel] = dict(ZERO_RESULT)
                logf.write(json.dumps({"bank": name, "eft": eft, "quarter": qlabel,
                                        "status": "no_consolidated_report"}, ensure_ascii=False) + "\n")
                continue
            try:
                cache_key = f"{eft}_{year}_{month:02d}"
                pdf_path = download_pdf(session, rapor_url, args.cache_dir, cache_key)
                result, status = extract_consumer_loans(pdf_path)
                bank_results[name][qlabel] = result
                logf.write(json.dumps({"bank": name, "eft": eft, "quarter": qlabel,
                                        "status": status}, ensure_ascii=False) + "\n")
                if status != "ok":
                    print(f"  NOTE {qlabel}: {status} -> missing fields default to 0")
            except Exception as e:
                print(f"  ERROR {qlabel}: {e} -> defaulting to 0")
                bank_results[name][qlabel] = dict(ZERO_RESULT)
                logf.write(json.dumps({"bank": name, "eft": eft, "quarter": qlabel,
                                        "status": f"error: {e}"}, ensure_ascii=False) + "\n")
            time.sleep(args.delay)

        # Save incrementally so a long run can be interrupted safely.
        wb = build_workbook(quarters, bank_results, bank_order[:bi])
        wb.save(args.output)

    logf.close()
    print(f"Done. Wrote {args.output}")


if __name__ == "__main__":
    main()
